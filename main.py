import openpyxl
import qrcode
from tkinter import Tk, Button, Label, filedialog, Frame, Canvas, ttk
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font
import os

def generar_qr(numero, nombre_archivo):
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(numero)
    qr.make(fit=True)

    imagen_qr = qr.make_image(fill_color="black", back_color="white")
    imagen_qr.save(nombre_archivo)

def leer_excel(nombre_archivo):
    try:
        workbook = openpyxl.load_workbook(nombre_archivo)
        sheet = workbook.active
        numeros_telefono = [str(cell.value) for cell in sheet['A'] if cell.value is not None]
        return numeros_telefono
    except Exception as e:
        print(f"Error al leer el archivo Excel: {e}")
        return []

def procesar_archivo():
    nombre_archivo_excel = filedialog.askopenfilename(filetypes=[("Archivos Excel", "*.xlsx;*.xls")])
    if not nombre_archivo_excel:
        resultado_var.config(text="No se seleccionó ningún archivo.")
        return

    numeros_telefono = leer_excel(nombre_archivo_excel)

    if not numeros_telefono:
        resultado_var.config(text="No se encontraron datos.")
        return

    resultados = []
    for i, numero in enumerate(numeros_telefono, start=1):
        nombre_archivo_qr = f"QR_{i}.png"
        generar_qr(numero, nombre_archivo_qr)

        resultados.append(f"x{nombre_archivo_qr}")
        resultados.append(numero)

    # Actualizar el archivo Excel original
    try:
        workbook = openpyxl.load_workbook(nombre_archivo_excel)
        sheet = workbook.active

        for row, resultado in enumerate(resultados, start=1):
            if resultado.startswith("x"):
                img = Image(resultado[1:])
                img.anchor = f"A{row}"
                sheet.add_image(img)
            else:
                sheet.cell(row=row, column=1, value=resultado)

        # Dar formato a la columna A
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column].width = adjusted_width

        # Dar formato específico a la columna A
        for row in sheet.iter_rows(min_row=1, max_col=1, max_row=sheet.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(size=28, bold=True)

        # Configurar el ancho de la columna A y el alto de todas las filas
        sheet.column_dimensions['A'].width = 41
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
            for cell in row:
                sheet.row_dimensions[cell.row].height = 160

        # Guardar el archivo Excel
        workbook.save(nombre_archivo_excel)

        # Eliminar las imágenes temporales de los QR después de actualizar el archivo Excel
        for i in range(1, len(resultados) + 1):
            nombre_archivo_qr = f"QR_{i}.png"
            if os.path.exists(nombre_archivo_qr):
                os.remove(nombre_archivo_qr)

        resultado_var.config(text=f"Resultados actualizados en {nombre_archivo_excel}. Imágenes temporales eliminadas.")
    except Exception as e:
        resultado_var.config(text=f"Error al actualizar el archivo Excel: {e}")

# Crear la interfaz gráfica
root = Tk()

root.title("ULTRAHOST-Generador de QR desde Excel- By IBZAN")

# Configurar la ventana
root.geometry("450x300")

# Tema
style = ttk.Style()
style.theme_use("clam")

# Etiqueta para mostrar el resultado
resultado_var = Label(root, text="", wraplength=380, justify="left")
resultado_var.pack(pady=10)

# Botón para seleccionar el archivo y procesarlo
boton_procesar = ttk.Button(root, text="Seleccionar Archivo Excel", command=procesar_archivo, width=30)
boton_procesar.pack(pady=10)

# Línea divisoria
linea_divisoria = Canvas(root, height=2, width=380, bg="black")
linea_divisoria.pack(pady=10)

# Texto de servicios
texto_servicios = Label(root, text="SERVICIOS DE HOSTING, DOMINIOS Y DISEÑO WEB", font=("Helvetica", 12, "bold"))
texto_servicios.pack()

# Texto de diseño de programas y aplicaciones
texto_diseno = Label(root, text="DISEÑO DE PROGRAMAS Y APLICACIONES", font=("Helvetica", 12, "bold"))
texto_diseno.pack(pady=10)

# Botones de servicios
frame_botones_servicios = Frame(root)
frame_botones_servicios.pack()

# Botón de WEB
boton_web = ttk.Button(frame_botones_servicios, text="WEB", command=lambda: web_button_click("https://www.ultrahost.uk"), width=20)
boton_web.pack(side="left", padx=10)

# Botón de WHATSAPP
boton_whatsapp = ttk.Button(frame_botones_servicios, text="WHATSAPP", command=lambda: web_button_click("https://wa.me/447418353168"), width=20)
boton_whatsapp.pack(side="right", padx=10)

def web_button_click(url):
    import webbrowser
    webbrowser.open_new(url)

# Iniciar el bucle principal de la interfaz gráfica
root.mainloop()
